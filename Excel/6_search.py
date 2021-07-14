from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active
for row in ws.iter_rows(min_row=2): #첫번째 줄 제외한 (제목부분 제외한) 다음줄  부터 
    # 번호, 영어, 수학
    if int(row[1].value) > 80:  #row[1]은 영어 점수
        print(row[0].value, "번 학생은 영어 천재") #번호

#영어를 컴퓨터로 바꾸기
for row in ws.iter_rows(max_row=1):  #최대가 첫번째줄까지
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")