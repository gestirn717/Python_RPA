from openpyxl import Workbook

wb = Workbook() 
ws = wb.create_sheet() #새로운 sheet 기본이름으로 생성 
ws.title = "Mysheet" #sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" #rgb형탸로 값을 넣어주면 탭 색상 변경

ws1 = wb.create_sheet("YourSheet")   #주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet",2)  #2번째 인덱스에 sheet생성

new_ws = wb["NewSheet"]    #Dict 형태로 sheet에 접근 / 시트 접근하기 

print(wb.sheetnames)#모든 sheet 이름 확인

#sheet복사
new_ws["A1"] = "Test"               #입력할 위치와 내용
target = wb.copy_worksheet(new_ws)  #target은 복사할 시트
target.title = "Copied Sheet"       #붙여넣기할 새로운 시트 


wb.save("sample.xlsx")